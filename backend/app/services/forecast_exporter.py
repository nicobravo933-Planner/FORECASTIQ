"""EXP-1a — Exportador analítico de resultados de forecast.

Genera un Excel con múltiples hojas:
  Hoja 1 — Resumen: dataset, modelo, métricas clave
  Hoja 2 — Predicciones: fecha, predicho, lower_ci, upper_ci
  Hoja 3 — Error mensual: fecha, real, predicho, error%, BIAS acum%
  Hoja 4 — Parámetros del modelo usado

Uso:
    exporter = ForecastExporter(job_result)
    buffer: BytesIO = exporter.generate_xlsx()
    csv_str: str    = exporter.generate_csv()
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import numpy as np

# openpyxl es requerido por pandas para .xlsx — ya está en pyproject.toml
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Paleta de colores ────────────────────────────────────────────────────────

_NAVY = "1F4E78"  # header principal
_HEADER_FG = "FFFFFF"  # texto blanco en headers
_GREEN_BG = "C6EFCE"  # semáforo verde (WAPE < 20%)
_YELLOW_BG = "FFEB9C"  # semáforo amarillo (WAPE 20-40%)
_RED_BG = "FFC7CE"  # semáforo rojo (WAPE > 40%)
_GREEN_FG = "276221"
_YELLOW_FG = "9C5700"
_RED_FG = "9C0006"
_ROW_ALT = "F2F7FF"  # fila alternada azul claro


class ForecastExporter:
    """Genera archivos de exportación analítica desde el resultado de un forecast job."""

    def __init__(self, job_result: dict[str, Any]) -> None:
        self.r = job_result  # dict tal como lo devuelve get_forecast_result()

    # ── Interfaz pública ──────────────────────────────────────────────────────

    def generate_xlsx(self) -> io.BytesIO:
        """Genera un Excel en memoria con todas las hojas analíticas."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl no está instalado — correr: uv add openpyxl")

        wb = Workbook()
        wb.remove(wb.active)  # quitar hoja vacía default

        self._sheet_resumen(wb)
        self._sheet_predicciones(wb)
        self._sheet_error_mensual(wb)
        self._sheet_parametros(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def generate_csv(self) -> str:
        """CSV tabular de predicciones (fecha, predicho, lower, upper)."""
        preds = self.r.get("predictions") or []
        if not preds:
            return "date,predicted,lower,upper\n"
        lines = ["date,predicted,lower,upper"]
        for p in preds:
            lines.append(
                f"{p.get('date', '')},"
                f"{_safe_float(p.get('predicted')) or ''},"
                f"{_safe_float(p.get('lower')) or ''},"
                f"{_safe_float(p.get('upper')) or ''}"
            )
        return "\n".join(lines)

    def generate_json(self) -> dict[str, Any]:
        """Retorna el resultado completo como dict serializable."""
        return dict(self.r)

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────

    def _sheet_resumen(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Resumen")
        self._set_col_widths(ws, [30, 38])

        row = 1
        self._merge_title(ws, row, 1, 2, "ForecastIQ — Resumen del Análisis")
        row += 1
        self._merge_title(
            ws,
            row,
            1,
            2,
            f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            navy=False,
        )
        row += 2

        model_labels = {
            "moving_average": "Promedio Móvil",
            "holt_winters": "Holt-Winters",
            "sarima": "SARIMA",
            "lightgbm": "LightGBM",
        }
        freq_labels = {"D": "Diaria", "W": "Semanal", "M": "Mensual", "Q": "Trimestral"}
        metrics = self.r.get("metrics") or {}

        info_rows = [
            ("Job ID", self.r.get("job_id", "—")),
            ("Dataset ID", self.r.get("dataset_id", "—")),
            (
                "Modelo usado",
                model_labels.get(
                    str(self.r.get("model_used", "")), str(self.r.get("model_used", "—"))
                ),
            ),
            (
                "Frecuencia",
                freq_labels.get(str(self.r.get("freq", "M")), str(self.r.get("freq", "—"))),
            ),
            ("Horizonte", f"{self.r.get('horizon', '—')} períodos"),
            (
                "Test periods",
                (
                    f"{self.r.get('test_periods', 0)} períodos"
                    if self.r.get("test_periods")
                    else "Hold-out automático 20%"
                ),
            ),
            ("Creado en", str(self.r.get("created_at", "—"))[:19]),
        ]

        self._header_row(ws, row, ["Parámetro", "Valor"])
        row += 1
        for i, (k, v) in enumerate(info_rows):
            self._data_row(ws, row, [k, str(v)], alt=(i % 2 == 0))
            row += 1

        row += 1  # espaciado entre tablas

        # Tabla de métricas con semáforo WAPE
        self._header_row(ws, row, ["Métrica", "Valor"])
        row += 1

        def _fmt_pct(v: Any) -> str:
            return "N/A" if v is None else f"{float(v) * 100:.2f}%"

        def _fmt_abs(v: Any) -> str:
            return "N/A" if v is None else f"{float(v):.4f}"

        metric_rows = [
            ("WAPE  (error relativo ponderado)", _fmt_pct(metrics.get("wape"))),
            ("MAE   (error absoluto promedio)", _fmt_abs(metrics.get("mae"))),
            ("BIAS  (sesgo sistemático)", _fmt_pct(metrics.get("bias"))),
            ("RMSE  (error cuadrático medio)", _fmt_abs(metrics.get("rmse"))),
            ("FVA   (valor añadido vs. Naive)", _fmt_pct(metrics.get("fva"))),
        ]
        for i, (k, v) in enumerate(metric_rows):
            cell_k = ws.cell(row=row, column=1, value=k)
            cell_v = ws.cell(row=row, column=2, value=v)
            cell_k.font = Font(name="Calibri", size=11)
            cell_v.font = Font(name="Calibri", size=11, bold=True)
            if i % 2 == 0:
                cell_k.fill = PatternFill("solid", fgColor=_ROW_ALT)
            # Semáforo de color en celda WAPE
            if k.startswith("WAPE") and metrics.get("wape") is not None:
                wape = float(metrics["wape"])
                if wape < 0.20:
                    cell_v.fill = PatternFill("solid", fgColor=_GREEN_BG)
                    cell_v.font = Font(name="Calibri", size=11, bold=True, color=_GREEN_FG)
                elif wape < 0.40:
                    cell_v.fill = PatternFill("solid", fgColor=_YELLOW_BG)
                    cell_v.font = Font(name="Calibri", size=11, bold=True, color=_YELLOW_FG)
                else:
                    cell_v.fill = PatternFill("solid", fgColor=_RED_BG)
                    cell_v.font = Font(name="Calibri", size=11, bold=True, color=_RED_FG)
            row += 1

    # ── Hoja 2: Predicciones ──────────────────────────────────────────────────

    def _sheet_predicciones(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Predicciones")
        self._set_col_widths(ws, [14, 16, 16, 16])
        self._header_row(ws, 1, ["Fecha", "Predicho", "Lower CI", "Upper CI"])

        for i, p in enumerate(self.r.get("predictions") or []):
            self._data_row(
                ws,
                i + 2,
                [
                    str(p.get("date", ""))[:10],
                    _safe_float(p.get("predicted")),
                    _safe_float(p.get("lower")),
                    _safe_float(p.get("upper")),
                ],
                alt=(i % 2 == 0),
            )

    # ── Hoja 3: Error mensual (solo si hay test_periods) ─────────────────────

    def _sheet_error_mensual(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Error mensual")
        self._set_col_widths(ws, [14, 16, 16, 14, 16])

        test_actual = self.r.get("test_actual") or []
        test_predicted = self.r.get("test_predicted") or []

        if not test_actual or not test_predicted:
            ws.cell(
                row=1,
                column=1,
                value="Sin período de test — configurar test_periods > 0 y re-correr.",
            )
            return

        self._header_row(ws, 1, ["Fecha", "Real", "Predicho", "Error %", "BIAS acum %"])

        sum_real = (
            sum(float(a.get("value", 0)) for a in test_actual if a.get("value") is not None) or 1.0
        )

        bias_acum = 0.0
        for i, (a, p) in enumerate(zip(test_actual, test_predicted, strict=False)):
            real = _safe_float(a.get("value"))
            pred = _safe_float(p.get("predicted"))
            error_pct: float | None = None
            if real and real != 0 and pred is not None:
                error_pct = round((pred - real) / abs(real) * 100, 2)
            if pred is not None and real is not None:
                bias_acum += (pred - real) / sum_real * 100

            row_vals: list[Any] = [
                str(a.get("date", ""))[:10],
                real,
                pred,
                f"{error_pct:+.1f}%" if error_pct is not None else "N/A",
                f"{bias_acum:+.1f}%",
            ]
            row_idx = i + 2
            for c, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row_idx, column=c, value=val)
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=_ROW_ALT)

            # Semáforo en error %
            if error_pct is not None:
                err_cell = ws.cell(row=row_idx, column=4)
                if abs(error_pct) < 20:
                    err_cell.fill = PatternFill("solid", fgColor=_GREEN_BG)
                    err_cell.font = Font(name="Calibri", size=11, color=_GREEN_FG)
                else:
                    err_cell.fill = PatternFill("solid", fgColor=_RED_BG)
                    err_cell.font = Font(name="Calibri", size=11, color=_RED_FG)

    # ── Hoja 4: Parámetros del modelo ─────────────────────────────────────────

    def _sheet_parametros(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Parámetros")
        self._set_col_widths(ws, [32, 28])

        model_used = str(self.r.get("model_used", "—"))
        model_params = self.r.get("model_params") or {}

        self._merge_title(ws, 1, 1, 2, f"Parámetros del modelo: {model_used}")
        self._header_row(ws, 2, ["Parámetro", "Valor"])

        if not model_params:
            ws.cell(row=3, column=1, value="Sin parámetros disponibles para este modelo.")
            return

        flat = _flatten_dict(model_params)
        for i, (k, v) in enumerate(flat):
            self._data_row(ws, i + 3, [k, v], alt=(i % 2 == 0))

    # ── Helpers de estilo ─────────────────────────────────────────────────────

    def _header_row(self, ws: Any, row: int, cols: list[str]) -> None:
        fill = PatternFill("solid", fgColor=_NAVY)
        font = Font(name="Calibri", size=11, bold=True, color=_HEADER_FG)
        border = Border(bottom=Side(style="thin", color="AAAAAA"))
        for c, val in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _data_row(self, ws: Any, row: int, vals: list[Any], alt: bool = False) -> None:
        fill = PatternFill("solid", fgColor=_ROW_ALT) if alt else None
        font = Font(name="Calibri", size=11)
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _merge_title(
        self, ws: Any, row: int, c1: int, c2: int, text: str, navy: bool = True
    ) -> None:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1, value=text)
        if navy:
            cell.fill = PatternFill("solid", fgColor=_NAVY)
            cell.font = Font(name="Calibri", size=13, bold=True, color=_HEADER_FG)
        else:
            cell.font = Font(name="Calibri", size=10, color="666666")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    @staticmethod
    def _set_col_widths(ws: Any, widths: list[float]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ───────────────────────────────────────────────────────────────────
# MS-B3: Export Excel multi-hoja para benchmark de múltiples series
# ───────────────────────────────────────────────────────────────────


def generate_multi_serie_xlsx(benchmark_result: dict[str, Any]) -> io.BytesIO:
    """
    Genera un Excel multi-hoja para el resultado de un benchmark multi-serie.

    Hojas:
      Resumen       — KPIs globales (n_series, modelos, train_end, test_periods, duración)
      Ganadores     — unique_id, best_model, WAPE%, BIAS%, score (una fila por entidad)
      Ranking       — WAPE promedio y victorias por modelo
      Predicciones  — todas las predicciones futuras (unique_id, ds, predicted)

    Args:
        benchmark_result: dict retornado por run_batch_benchmark() o el endpoint /benchmark-dataset.

    Returns:
        BytesIO con el archivo .xlsx listo para StreamingResponse.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    navy = "1F4E78"  # noqa: N806
    green = "C6EFCE"  # noqa: N806
    orange = "FFEB9C"  # noqa: N806
    red = "FFC7CE"  # noqa: N806
    white = "FFFFFF"  # noqa: N806

    wb = Workbook()

    def _header_row(ws: Any, row: int, cols: list[str]) -> None:
        for c, name in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.font = Font(name="Calibri", size=11, bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18

    def _data_row(ws: Any, row: int, vals: list[Any], alt: bool = False) -> None:
        fill_color = "F2F2F2" if alt else white
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = Font(name="Calibri", size=11)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _wape_color(wape_pct: float | None) -> str:
        if wape_pct is None:
            return white
        if wape_pct > 35:
            return red
        if wape_pct > 20:
            return orange
        return green

    # ───────────────────────────────────────
    # Hoja 1: Resumen
    # ───────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.column_dimensions["A"].width = 28
    ws_resumen.column_dimensions["B"].width = 36

    title_cell = ws_resumen.cell(row=1, column=1, value="ForecastIQ — Benchmark Multi-serie")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=white)
    title_cell.fill = PatternFill("solid", fgColor=navy)
    ws_resumen.merge_cells("A1:B1")
    ws_resumen.row_dimensions[1].height = 24

    kpis = [
        ("Series analizadas", benchmark_result.get("n_series", "")),
        ("Períodos de test", benchmark_result.get("test_periods", "sin test")),
        ("Train end", benchmark_result.get("train_end", "")),
        ("Horizonte futuro", benchmark_result.get("horizon", "")),
        ("Frecuencia", benchmark_result.get("freq", "")),
        ("Modelos comparados", ", ".join(benchmark_result.get("models_used", []))),
        ("Duración (s)", round(float(benchmark_result.get("duration_s", 0)), 2)),
        ("Exportado", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r, (label, value) in enumerate(kpis, start=3):
        ws_resumen.cell(row=r, column=1, value=label).font = Font(
            name="Calibri", size=11, bold=True
        )
        ws_resumen.cell(row=r, column=2, value=value).font = Font(name="Calibri", size=11)

    # Series omitidas (si las hay)
    skipped = benchmark_result.get("series_skipped", [])
    if skipped:
        r_skip = len(kpis) + 5
        ws_resumen.cell(
            row=r_skip, column=1, value="Series omitidas (datos insuficientes)"
        ).font = Font(name="Calibri", size=11, bold=True, color="C00000")
        ws_resumen.cell(
            row=r_skip,
            column=2,
            value=", ".join(skipped[:20])
            + (f" y {len(skipped) - 20} más" if len(skipped) > 20 else ""),
        ).font = Font(name="Calibri", size=11, color="C00000")

    # ───────────────────────────────────────
    # Hoja 2: Ganadores por entidad
    # ───────────────────────────────────────
    ws_ganadores = wb.create_sheet("Ganadores")
    _header_row(ws_ganadores, 1, ["Entidad", "Mejor modelo", "WAPE %", "BIAS %", "Score"])
    ws_ganadores.column_dimensions["A"].width = 30
    ws_ganadores.column_dimensions["B"].width = 18
    ws_ganadores.column_dimensions["C"].width = 12
    ws_ganadores.column_dimensions["D"].width = 12
    ws_ganadores.column_dimensions["E"].width = 12

    best_models: list[dict[str, Any]] = benchmark_result.get("best_models", [])
    # Ordenar por WAPE ascendente (mejor primero)
    best_models_sorted = sorted(best_models, key=lambda r: r.get("wape") or 1.0)

    for r_idx, bm in enumerate(best_models_sorted, start=2):
        wape_pct = round(float(bm["wape"]) * 100, 2) if bm.get("wape") is not None else None
        bias_pct = round(float(bm["bias"]), 2) if bm.get("bias") is not None else None
        score = round(float(bm["score"]), 4) if bm.get("score") is not None else None
        vals = [bm["unique_id"], bm["best_model"], wape_pct, bias_pct, score]
        _data_row(ws_ganadores, r_idx, vals, alt=(r_idx % 2 == 0))
        # Color semáforo WAPE
        wape_cell = ws_ganadores.cell(row=r_idx, column=3)
        if wape_pct is not None:
            wape_cell.fill = PatternFill("solid", fgColor=_wape_color(wape_pct))

    # ───────────────────────────────────────
    # Hoja 3: Ranking de modelos
    # ───────────────────────────────────────
    ws_ranking = wb.create_sheet("Ranking")
    _header_row(ws_ranking, 1, ["Modelo", "WAPE medio %", "BIAS medio %", "Victorias"])
    ws_ranking.column_dimensions["A"].width = 18
    ws_ranking.column_dimensions["B"].width = 16
    ws_ranking.column_dimensions["C"].width = 16
    ws_ranking.column_dimensions["D"].width = 14

    model_ranking: list[dict[str, Any]] = benchmark_result.get("model_ranking", [])
    for r_idx, mr in enumerate(model_ranking, start=2):
        wape_pct_r = round(float(mr["wape_mean"]) * 100, 2)
        bias_pct_r = round(float(mr["bias_mean"]), 2)
        vals = [mr["model"], wape_pct_r, bias_pct_r, mr["n_wins"]]
        _data_row(ws_ranking, r_idx, vals, alt=(r_idx % 2 == 0))
        ws_ranking.cell(row=r_idx, column=2).fill = PatternFill(
            "solid", fgColor=_wape_color(wape_pct_r)
        )

    # ───────────────────────────────────────
    # Hoja 4: Predicciones futuras (todas las series)
    # ───────────────────────────────────────
    ws_preds = wb.create_sheet("Predicciones")
    _header_row(ws_preds, 1, ["Entidad", "Fecha", "Predicción"])
    ws_preds.column_dimensions["A"].width = 30
    ws_preds.column_dimensions["B"].width = 16
    ws_preds.column_dimensions["C"].width = 18

    predictions: list[dict[str, Any]] = benchmark_result.get("predictions", [])
    # Ordenar por entidad + fecha
    predictions_sorted = sorted(
        predictions, key=lambda p: (str(p.get("unique_id", "")), str(p.get("ds", "")))
    )
    for r_idx, pred in enumerate(predictions_sorted, start=2):
        uid = str(pred.get("unique_id", ""))
        ds = str(pred.get("ds", ""))
        val = (
            round(float(pred.get("predicted", 0)), 4) if pred.get("predicted") is not None else None
        )
        _data_row(ws_preds, r_idx, [uid, ds, val], alt=(r_idx % 2 == 0))

    # ───────────────────────────────────────
    # Hoja 5: Test vs Real (P1 — S&OP validation sheet)
    # ───────────────────────────────────────
    test_vs_real: list[dict[str, Any]] = benchmark_result.get("test_vs_real", []) or []
    ws_test = wb.create_sheet("Test vs Real")
    _header_row(ws_test, 1, ["Entidad", "Fecha", "Real", "Predicho", "Error %", "BIAS acum %"])
    ws_test.column_dimensions["A"].width = 30
    ws_test.column_dimensions["B"].width = 14
    ws_test.column_dimensions["C"].width = 14
    ws_test.column_dimensions["D"].width = 14
    ws_test.column_dimensions["E"].width = 12
    ws_test.column_dimensions["F"].width = 14

    if not test_vs_real:
        ws_test.cell(
            row=2,
            column=1,
            value="Sin datos de test — configurar train_end antes del último período del dataset.",
        ).font = Font(name="Calibri", size=11, italic=True, color="888888")
    else:
        tv_sorted = sorted(
            test_vs_real, key=lambda r: (str(r.get("unique_id", "")), str(r.get("ds", "")))
        )
        for r_idx, tv_row in enumerate(tv_sorted, start=2):
            err_pct_val = tv_row.get("error_pct")
            bias_val = tv_row.get("bias_acum_pct")
            vals = [
                str(tv_row.get("unique_id", "")),
                str(tv_row.get("ds", ""))[:10],
                tv_row.get("real"),
                tv_row.get("predicted"),
                f"{err_pct_val:+.1f}%" if err_pct_val is not None else "N/A",
                f"{bias_val:+.1f}%" if bias_val is not None else "N/A",
            ]
            _data_row(ws_test, r_idx, vals, alt=(r_idx % 2 == 0))
            # Semáforo en Error %: verde < 20%, naranja 20-35%, rojo > 35%
            if err_pct_val is not None:
                err_cell = ws_test.cell(row=r_idx, column=5)
                abs_err = abs(err_pct_val)
                if abs_err < 20:
                    err_cell.fill = PatternFill("solid", fgColor=green)
                    err_cell.font = Font(name="Calibri", size=11, color="276221")
                elif abs_err < 35:
                    err_cell.fill = PatternFill("solid", fgColor=orange)
                    err_cell.font = Font(name="Calibri", size=11, color="9C5700")
                else:
                    err_cell.fill = PatternFill("solid", fgColor=red)
                    err_cell.font = Font(name="Calibri", size=11, color="9C0006")

    # Guardar en buffer y retornar
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Helpers globales ──────────────────────────────────────────────────────────


def _safe_float(v: Any) -> float | None:
    """Convierte a float o None si no es convertible o es NaN/Inf."""
    try:
        f = float(v)
        return None if (np.isnan(f) or np.isinf(f)) else round(f, 4)
    except (TypeError, ValueError):
        return None


def _flatten_dict(d: dict[str, Any], prefix: str = "") -> list[tuple[str, str]]:
    """Aplana dicts anidados para la hoja Parámetros (e.g. LightGBM best_params)."""
    items: list[tuple[str, str]] = []
    for k, v in d.items():
        full_key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            items.extend(_flatten_dict(v, full_key))
        else:
            items.append((full_key, str(v)))
    return items


# ── EXP-1c: Exportador de benchmark multi-modelo ─────────────────────────────

# Color de acento para el modelo ganador
_WINNER_BG = "D9E8FF"  # azul claro
_WINNER_FG = "1F4E78"  # navy


class BenchmarkExporter:
    """EXP-1c — Genera un Excel analítico del resultado del benchmark multi-modelo.

    Estructura del Excel:
      Hoja 1 — Comparación: tabla de todos los modelos con métricas + semáforos
      Hoja 2 — Conclusión: ganador, FVA, texto educativo y guía de interpretación
      Hoja N — Pred <modelo>: predicciones futuras de cada modelo que las tenga

    Recibe el BenchmarkResult serializado como dict (igual que lo devuelve
    el endpoint POST /api/forecast/benchmark al frontend).
    """

    def __init__(self, benchmark: dict[str, Any]) -> None:
        self.b = benchmark

    def generate_xlsx(self) -> io.BytesIO:
        """Genera el Excel en memoria y retorna un BytesIO listo para StreamingResponse."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl no está instalado — correr: uv add openpyxl")

        wb = Workbook()
        wb.remove(wb.active)  # quitar hoja vacía default

        self._sheet_comparacion(wb)
        self._sheet_conclusion(wb)
        self._sheets_predicciones(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Hoja 1: Comparación de modelos ───────────────────────────────────────

    def _sheet_comparacion(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Comparación")
        # Modelo | WAPE | MAE | BIAS | RMSE | FVA | Estado
        self._set_col_widths(ws, [28, 12, 14, 12, 14, 12, 16])

        row = 1
        self._merge_title(ws, row, 1, 7, "ForecastIQ — Benchmark Multi-Modelo")
        row += 1
        freq_labels = {"D": "Diaria", "W": "Semanal", "M": "Mensual", "Q": "Trimestral"}
        freq_str = freq_labels.get(str(self.b.get("freq", "M")), str(self.b.get("freq", "M")))
        meta = (
            f"Dataset: {str(self.b.get('dataset_id', ''))[:20]}  |  "
            f"Frecuencia: {freq_str}  |  "
            f"Horizonte: {self.b.get('horizon', '—')}p  |  "
            f"Test: {self.b.get('test_periods', 0)}p  |  "
            f"Obs.: {self.b.get('n_obs', '—')}  |  "
            f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        )
        self._merge_title(ws, row, 1, 7, meta, navy=False)
        row += 2

        self._header_row(ws, row, ["Modelo", "WAPE", "MAE", "BIAS", "RMSE", "FVA", "Estado"])
        row += 1

        winner_id = self.b.get("winner")
        models: list[dict[str, Any]] = self.b.get("models") or []

        for i, m in enumerate(models):
            is_winner = m.get("model") == winner_id and not m.get("is_baseline")
            is_naive = bool(m.get("is_baseline"))
            has_error = bool(m.get("error"))
            wape_val = _safe_float(m.get("wape"))
            bias_f = _safe_float(m.get("bias"))
            mae_f = _safe_float(m.get("mae"))
            rmse_f = _safe_float(m.get("rmse"))
            fva_f = _safe_float(m.get("fva"))

            estado = (
                f"Error: {str(m.get('error', ''))[:30]}"
                if has_error
                else "✓ Ganador"
                if is_winner
                else "Baseline"
                if is_naive
                else "OK"
            )

            row_vals: list[Any] = [
                m.get("label", m.get("model", "—")),
                f"{wape_val * 100:.2f}%" if wape_val is not None else "N/A",
                f"{mae_f:.4f}" if mae_f is not None else "N/A",
                f"{(bias_f or 0) * 100:+.2f}%" if bias_f is not None else "N/A",
                f"{rmse_f:.4f}" if rmse_f is not None else "N/A",
                f"{fva_f:+.1f}%" if fva_f is not None else "—",
                estado,
            ]

            alt = i % 2 == 0
            for c, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if alt:
                    cell.fill = PatternFill("solid", fgColor=_ROW_ALT)

            # Resaltar fila ganadora en azul suave (sobreescribe el alt fill)
            if is_winner:
                for c in range(1, 8):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = PatternFill("solid", fgColor=_WINNER_BG)
                    cell.font = Font(name="Calibri", size=11, bold=True, color=_WINNER_FG)

            # Semáforo WAPE columna 2 (solo no-ganadores para que el azul domine)
            if wape_val is not None and not is_winner:
                wape_cell = ws.cell(row=row, column=2)
                if wape_val < 0.20:
                    wape_cell.fill = PatternFill("solid", fgColor=_GREEN_BG)
                    wape_cell.font = Font(name="Calibri", size=11, bold=True, color=_GREEN_FG)
                elif wape_val < 0.40:
                    wape_cell.fill = PatternFill("solid", fgColor=_YELLOW_BG)
                    wape_cell.font = Font(name="Calibri", size=11, bold=True, color=_YELLOW_FG)
                else:
                    wape_cell.fill = PatternFill("solid", fgColor=_RED_BG)
                    wape_cell.font = Font(name="Calibri", size=11, bold=True, color=_RED_FG)

            # Semáforo FVA columna 6 (verde = modelo agrega valor)
            if fva_f is not None and not is_winner:
                fva_cell = ws.cell(row=row, column=6)
                if fva_f > 0:
                    fva_cell.fill = PatternFill("solid", fgColor=_GREEN_BG)
                    fva_cell.font = Font(name="Calibri", size=11, color=_GREEN_FG)
                elif fva_f < 0:
                    fva_cell.fill = PatternFill("solid", fgColor=_RED_BG)
                    fva_cell.font = Font(name="Calibri", size=11, color=_RED_FG)

            row += 1

    # ── Hoja 2: Conclusión y guía ─────────────────────────────────────────────

    def _sheet_conclusion(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Conclusión")
        self._set_col_widths(ws, [28, 52])

        row = 1
        self._merge_title(ws, row, 1, 2, "Conclusión del Benchmark")
        row += 2

        winner_label = self.b.get("winner_label") or "—"
        naive_wape = _safe_float(self.b.get("naive_wape"))
        conclusion = str(self.b.get("conclusion") or "—")

        info: list[tuple[str, str]] = [
            ("Modelo ganador", winner_label),
            ("WAPE Seasonal Naive", f"{naive_wape * 100:.2f}%" if naive_wape else "N/A"),
            ("Conclusión automática", conclusion),
            ("Corrido en", str(self.b.get("run_at", ""))[:19]),
        ]

        self._header_row(ws, row, ["Campo", "Valor"])
        row += 1
        for i, (k, v) in enumerate(info):
            cell_k = ws.cell(row=row, column=1, value=k)
            cell_v = ws.cell(row=row, column=2, value=v)
            cell_k.font = Font(name="Calibri", size=11)
            cell_v.font = Font(name="Calibri", size=11)
            cell_v.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 28
            if i % 2 == 0:
                cell_k.fill = PatternFill("solid", fgColor=_ROW_ALT)
                cell_v.fill = PatternFill("solid", fgColor=_ROW_ALT)
            if k == "Modelo ganador":
                cell_v.fill = PatternFill("solid", fgColor=_WINNER_BG)
                cell_v.font = Font(name="Calibri", size=11, bold=True, color=_WINNER_FG)
            row += 1

        # Guía de interpretación (Vandeputt)
        row += 2
        self._merge_title(ws, row, 1, 2, "Guía de interpretación (Vandeputt)", navy=False)
        row += 1
        guia: list[tuple[str, str]] = [
            (
                "WAPE",
                "Error relativo ponderado. < 20% excelente, 20-40% aceptable, > 40% revisar datos.",
            ),
            (
                "BIAS",
                "Sesgo sistemático. + = sobreestimás (riesgo sobrestock). - = subestimás (riesgo quiebre).",
            ),
            (
                "FVA",
                "Forecast Value Added. + = tu modelo supera al Seasonal Naive. - = usá el Naive directamente.",
            ),
            (
                "Naive",
                "Seasonal Naive = copiar el valor del mismo período del año anterior. Baseline mínimo a superar.",
            ),
            ("Ganador", "Modelo con menor WAPE entre los no-Naive con run exitoso."),
        ]
        for k, v in guia:
            cell_k = ws.cell(row=row, column=1, value=k)
            cell_v = ws.cell(row=row, column=2, value=v)
            cell_k.font = Font(name="Calibri", size=10, bold=True)
            cell_v.font = Font(name="Calibri", size=10)
            cell_v.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1

    # ── Hojas de predicciones por modelo ──────────────────────────────────────

    def _sheets_predicciones(self, wb: Workbook) -> None:
        """Una hoja por modelo que tenga predicciones futuras disponibles."""
        model_predictions: dict[str, list[dict[str, Any]]] = self.b.get("model_predictions") or {}
        model_labels: dict[str, str] = {
            m.get("model", ""): m.get("label", m.get("model", ""))
            for m in (self.b.get("models") or [])
        }
        winner_id = self.b.get("winner")

        for model_id, preds in model_predictions.items():
            if not preds:
                continue
            label = model_labels.get(model_id, model_id)
            sheet_name = f"Pred {label}"[:31]  # Excel: máx 31 chars por nombre de hoja
            ws = wb.create_sheet(sheet_name)
            self._set_col_widths(ws, [14, 16, 16, 16])

            titulo = f"Predicciones — {label}"
            if model_id == winner_id:
                titulo += " ✓ GANADOR"
            self._merge_title(ws, 1, 1, 4, titulo)
            self._header_row(ws, 2, ["Fecha", "Predicho", "Lower CI", "Upper CI"])

            for i, p in enumerate(preds):
                self._data_row(
                    ws,
                    i + 3,
                    [
                        str(p.get("date", ""))[:10],
                        _safe_float(p.get("predicted")),
                        _safe_float(p.get("lower")),
                        _safe_float(p.get("upper")),
                    ],
                    alt=(i % 2 == 0),
                )

    # ── Helpers de estilo ─────────────────────────────────────────────────────

    def _header_row(self, ws: Any, row: int, cols: list[str]) -> None:
        fill = PatternFill("solid", fgColor=_NAVY)
        font = Font(name="Calibri", size=11, bold=True, color=_HEADER_FG)
        border = Border(bottom=Side(style="thin", color="AAAAAA"))
        for c, val in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _data_row(self, ws: Any, row: int, vals: list[Any], alt: bool = False) -> None:
        fill = PatternFill("solid", fgColor=_ROW_ALT) if alt else None
        font = Font(name="Calibri", size=11)
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _merge_title(
        self, ws: Any, row: int, c1: int, c2: int, text: str, navy: bool = True
    ) -> None:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1, value=text)
        if navy:
            cell.fill = PatternFill("solid", fgColor=_NAVY)
            cell.font = Font(name="Calibri", size=13, bold=True, color=_HEADER_FG)
        else:
            cell.font = Font(name="Calibri", size=10, color="666666")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    @staticmethod
    def _set_col_widths(ws: Any, widths: list[float]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
